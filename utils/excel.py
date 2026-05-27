from openpyxl import Workbook, load_workbook
from utils.db import Excel, excel_fs
import os
import tempfile

def create_empty_excel(email, project_name):

    file_path = f"{project_name}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Certificate Data"

    ws.append(["Name", "College", "Year", "Reg No"])

    wb.save(file_path)

    try:
        with open(file_path, "rb") as f:
            file_id = excel_fs.put(
                f.read(),
                filename=f"{project_name}.xlsx",
                metadata={
                    "email": email,
                    "project_name": project_name
                }
            )

        Excel.update_one(
            {"email": email, "project_name": project_name},
            {"$set": {"excel_file_id": file_id}},
            upsert=True
        )

    finally:
        if os.path.exists(file_path):
            os.remove(file_path)

def add_data_to_excel(email, project_name, new_data):
    record = Excel.find_one({
        "email": email,
        "project_name": project_name
    })

    if not record:
        return

    old_file_id = record["excel_file_id"]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        temp_path = tmp.name

    try:
        file_data = excel_fs.get(old_file_id)

        with open(temp_path, "wb") as f:
            f.write(file_data.read())

        wb = load_workbook(temp_path)
        ws = wb.active

        reg_no = new_data.get("certId") or new_data.get("cert_id")

        existing_reg_nos = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            existing_reg_nos.add(str(row[3]))

        if str(reg_no) in existing_reg_nos:
            print("⚠ Already exists in Excel → Skipping")
            return

        ws.append([
            new_data.get("name"),
            new_data.get("college"),
            new_data.get("year"),
            reg_no
        ])

        wb.save(temp_path)

        with open(temp_path, "rb") as f:
            new_file_id = excel_fs.put(
                f.read(),
                filename=f"{project_name}.xlsx",
                metadata={
                    "email": email,
                    "project_name": project_name
                }
            )

        Excel.update_one(
            {"_id": record["_id"]},
            {"$set": {"excel_file_id": new_file_id}}
        )

        excel_fs.delete(old_file_id)

    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)